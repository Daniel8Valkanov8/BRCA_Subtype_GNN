# -*- coding: utf-8 -*-
"""
Генерира Excel таблица с всички 196 гена от модела best_model_tcga_196genes_09.06.pt.

Tab 1 «Всички гени (PAM50)»   : всички гени; PAM50 гените са преместени най-отдолу и
                                оцветени в червено.
Tab 2 «Гени по субтип»        : всички гени, цветово кодирани по BRCA субтип, който
                                помагат да се разграничи (легенда вдясно).
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import torch
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1) Извличане на гените от модела / данните
# ---------------------------------------------------------------------------
EXPR_CSV = 'data/tcga_expression_198genes.csv'
PPI_TSV = 'data/string_ppi_196genes.tsv'
MODEL_PATH = 'best_model_tcga_196genes_09.06.pt'

genes = list(pd.read_csv(EXPR_CSV, index_col=0).index)   # реда = реда на нодовете в графа
print(f"Гени от данните на модела: {len(genes)}")

# STRING PPI степен (брой ребра) за всеки ген -> свързан vs изолиран връх
from collections import Counter
_ppi = pd.read_csv(PPI_TSV, sep='\t')
degree = Counter()
for a, b in zip(_ppi['node1'], _ppi['node2']):
    degree[a] += 1
    degree[b] += 1
n_isolated = sum(1 for g in genes if degree.get(g, 0) == 0)
print(f"STRING ребра: {len(_ppi)} | изолирани върхове: {n_isolated}/{len(genes)}")

# Проверка, че checkpoint-ът е съвместим (брой нодове)
try:
    state = torch.load(MODEL_PATH, map_location='cpu', weights_only=True)
    print(f"Checkpoint зареден: {MODEL_PATH}  ({len(state)} тензора)")
except Exception as e:
    print(f"(checkpoint info skip: {e})")

# ---------------------------------------------------------------------------
# 2) Официален PAM50 списък (Parker et al., 2009, JCO)
# ---------------------------------------------------------------------------
PAM50 = {
    "ACTR3B","ANLN","BAG1","BCL2","BIRC5","BLVRA","CCNB1","CCNE1","CDC20","CDC6",
    "CDH3","CENPF","CEP55","CXXC5","EGFR","ERBB2","ESR1","EXO1","FGFR4","FOXA1",
    "FOXC1","GPR160","GRB7","HMGB3","KIF2C","KNTC2","LAMA2","MAPT","MDM2","MELK",
    "MIA","MKI67","MLPH","MMP11","MYBL2","MYC","NAT1","NDC80","ORC6L","PHGDH",
    "PTTG1","RRM2","SFRP1","SLC39A6","TMEM45B","TYMS","UBE2C","UBE2T","VEGFA","PGR",
}
pam50_present = [g for g in genes if g in PAM50]
print(f"PAM50 гени в модела: {len(pam50_present)} -> {pam50_present}")

# ---------------------------------------------------------------------------
# 3) Субтип-маркери (документирани топ гени по подтип)
#    Източник: BRCA_Gene_Signatures_Analysis.txt + check_pam50.py анотации.
#    primary = основният подтип, който гена помага да се разграничи.
# ---------------------------------------------------------------------------
SUBTYPE_GENES = {
    "BRCA_LumA":   ["ESR1","FOXA1","GATA3","PGR","CDKN1A","CDKN1B","TFF1","GREB1","AREG","MUC1",
                    "BAG1","BCL2","BLVRA","CXXC5","MIA","NAT1","SFRP1","SLC39A6","MAPT","MLPH","FGFR4","GPR160"],
    "BRCA_LumB":   ["GRB2","CDK2","CCNE2","CCNB1","E2F1","MYBL2","UBE2C","PTTG1","CCNE1","BIRC5","TMEM45B"],
    "BRCA_Her2":   ["ERBB2","GRB7","PIK3CA","JAK2","EGF","MDM2","MMP11","VEGFA","MET","STAT3"],
    "BRCA_Basal":  ["TP53","BRCA1","MYC","KRAS","BRAF","KIT","KRT5","KRT14","NOTCH1","PARP1","MTOR",
                    "EGFR","FOXC1","CDH3","MELK","PHGDH","CEP55","ANLN","UBE2T","ACTR3B","MKI67"],
    "BRCA_Normal": ["EPCAM","CDH1","ANXA5","LAMA2"],
}
# Първична принадлежност (ако ген е в няколко списъка → взима се по-специфичният подтип).
# Ред на приоритет: Basal > Her2 > LumB > Normal > LumA  (по-редките/специфичните печелят)
PRIORITY = ["BRCA_Basal","BRCA_Her2","BRCA_LumB","BRCA_Normal","BRCA_LumA"]
gene_subtype = {}
for st in PRIORITY:
    for g in SUBTYPE_GENES.get(st, []):
        if g in genes and g not in gene_subtype:
            gene_subtype[g] = st

# ---------------------------------------------------------------------------
# Цветова схема
# ---------------------------------------------------------------------------
SUBTYPE_COLORS = {
    "BRCA_Basal":  "C00000",  # тъмно червено
    "BRCA_Her2":   "7030A0",  # лилаво
    "BRCA_LumA":   "2E75B6",  # синьо
    "BRCA_LumB":   "00B050",  # зелено
    "BRCA_Normal": "ED7D31",  # оранжево
    None:          "BFBFBF",  # сиво — без специфична субтип-роля
}
SUBTYPE_LABEL = {
    "BRCA_Basal":  "Basal",
    "BRCA_Her2":   "Her2",
    "BRCA_LumA":   "Luminal A",
    "BRCA_LumB":   "Luminal B",
    "BRCA_Normal": "Normal-like",
    None:          "Общ / без специфична роля",
}

thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="1F1F1F")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER

wb = Workbook()

# ===========================================================================
# TAB 1 — Всички гени, PAM50 най-отдолу, червени
# ===========================================================================
ws1 = wb.active
ws1.title = "Всички гени (PAM50)"

ws1.cell(1, 1, "Всички 196 гена от модела best_model_tcga_196genes_09.06.pt").font = Font(bold=True, size=13)
ws1.cell(2, 1, "Не-PAM50 гените са горе; PAM50 гените са най-отдолу (червено). Колоната STRING показва свързани (с ребра) vs изолирани върхове.").font = Font(italic=True, size=10, color="666666")

hdr = 4
ws1.cell(hdr, 1, "№"); ws1.cell(hdr, 2, "Ген"); ws1.cell(hdr, 3, "PAM50?")
ws1.cell(hdr, 4, "Тип"); ws1.cell(hdr, 5, "STRING"); ws1.cell(hdr, 6, "Степен")
style_header(ws1, hdr, 6)

non_pam = [g for g in genes if g not in PAM50]
ordered = non_pam + pam50_present   # PAM50 най-отдолу
RED_FILL = PatternFill("solid", fgColor="FFC7CE")        # PAM50 свързан
RED_FONT = Font(color="9C0006", bold=True)
ISO_FILL = PatternFill("solid", fgColor="FFE699")        # изолиран връх (жълто)
ISO_FONT = Font(color="9C6500", bold=True)

r = hdr + 1
for i, g in enumerate(ordered, 1):
    is_pam = g in PAM50
    deg = degree.get(g, 0)
    isolated = deg == 0
    ws1.cell(r, 1, i).alignment = CENTER
    ws1.cell(r, 2, g)
    ws1.cell(r, 3, "ДА" if is_pam else "—").alignment = CENTER
    ws1.cell(r, 4, "PAM50" if is_pam else "Разширен панел")
    ws1.cell(r, 5, "Изолиран" if isolated else "Свързан").alignment = CENTER
    ws1.cell(r, 6, deg).alignment = CENTER
    for c in range(1, 7):
        cell = ws1.cell(r, c)
        cell.border = BORDER
        if is_pam:
            cell.fill = RED_FILL
            cell.font = RED_FONT
    # изолираните върхове (колони STRING/Степен) се маркират в жълто за всички гени
    if isolated:
        ws1.cell(r, 5).fill = ISO_FILL
        ws1.cell(r, 5).font = ISO_FONT
        ws1.cell(r, 6).fill = ISO_FILL
        ws1.cell(r, 6).font = ISO_FONT
    r += 1

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 9
ws1.freeze_panes = "A5"

# Обобщение под таблицата
sumrow = r + 1
pam_iso = [g for g in pam50_present if degree.get(g, 0) == 0]
pam_con = [g for g in pam50_present if degree.get(g, 0) > 0]
ws1.cell(sumrow, 2, f"PAM50 общо: {len(pam50_present)}  |  Свързани: {len(pam_con)}  |  Изолирани: {len(pam_iso)}").font = Font(bold=True)
ws1.cell(sumrow + 1, 2, "Изолирани PAM50 върхове: " + ", ".join(pam_iso)).font = Font(color="9C6500")

# ===========================================================================
# TAB 2 — Всички гени, цветово кодирани по субтип
# ===========================================================================
ws2 = wb.create_sheet("Гени по субтип")
ws2.cell(1, 1, "Гени по BRCA субтип, който помагат да се разграничи").font = Font(bold=True, size=13)
ws2.cell(2, 1, "Всеки субтип има свой цвят (виж легендата вдясно). Сивите гени нямат специфична субтип-роля.").font = Font(italic=True, size=10, color="666666")

hdr = 4
ws2.cell(hdr, 1, "№"); ws2.cell(hdr, 2, "Ген"); ws2.cell(hdr, 3, "Субтип"); ws2.cell(hdr, 4, "PAM50?")
style_header(ws2, hdr, 4)

# подреждаме по субтип (групирано), после общи
order_map = {st: i for i, st in enumerate(["BRCA_Basal","BRCA_Her2","BRCA_LumB","BRCA_LumA","BRCA_Normal"])}
def sort_key(g):
    st = gene_subtype.get(g)
    return (order_map.get(st, 99), g)
genes_sorted = sorted(genes, key=sort_key)

r = hdr + 1
for i, g in enumerate(genes_sorted, 1):
    st = gene_subtype.get(g)
    color = SUBTYPE_COLORS[st]
    ws2.cell(r, 1, i).alignment = CENTER
    ws2.cell(r, 2, g)
    ws2.cell(r, 3, SUBTYPE_LABEL[st])
    ws2.cell(r, 4, "ДА" if g in PAM50 else "—").alignment = CENTER
    font_color = "FFFFFF" if st is not None else "000000"
    for c in range(1, 5):
        cell = ws2.cell(r, c)
        cell.border = BORDER
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color=font_color, bold=(c == 2))
    r += 1

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 10
ws2.freeze_panes = "A5"

# Легенда вдясно
lr = hdr
ws2.cell(lr, 6, "ЛЕГЕНДА — цвят по субтип").font = Font(bold=True, size=11)
lr += 1
counts = {}
for g in genes:
    counts[gene_subtype.get(g)] = counts.get(gene_subtype.get(g), 0) + 1
for st in ["BRCA_Basal","BRCA_Her2","BRCA_LumA","BRCA_LumB","BRCA_Normal", None]:
    cell = ws2.cell(lr, 6, SUBTYPE_LABEL[st])
    cell.fill = PatternFill("solid", fgColor=SUBTYPE_COLORS[st])
    cell.font = Font(color="FFFFFF" if st is not None else "000000", bold=True)
    cell.border = BORDER
    cell.alignment = LEFT
    ws2.cell(lr, 7, f"{counts.get(st,0)} гена").alignment = LEFT
    lr += 1
ws2.column_dimensions['F'].width = 26
ws2.column_dimensions['G'].width = 12

# ===========================================================================
# TAB 3 — Гени по субтип СПОРЕД МОДЕЛА (градиентна saliency)
# TAB 4 — Сравнение литература vs модел
# Изисква gene_importance_by_class.csv (от model_gene_importance.py)
# ===========================================================================
import os
IMP_CSV = "gene_importance_by_class.csv"
model_assign = {}
if os.path.exists(IMP_CSV):
    imp = pd.read_csv(IMP_CSV, index_col=0)
    cls_cols = ["BRCA_Basal","BRCA_Her2","BRCA_LumA","BRCA_LumB","BRCA_Normal"]
    cls_cols = [c for c in cls_cols if c in imp.columns]
    model_assign = imp['assigned_subtype'].to_dict()

    ws3 = wb.create_sheet("Гени по субтип (модел)")
    ws3.cell(1, 1, "Гени по субтип СПОРЕД МОДЕЛА — градиентна saliency на best_model_tcga_196genes_09.06.pt").font = Font(bold=True, size=13)
    ws3.cell(2, 1, "Назначеният субтип = класът, чийто логит е най-чувствителен към експресията на гена (нормализирано по клас). Това НЕ е литература — извлечено е от теглата на модела.").font = Font(italic=True, size=10, color="666666")

    hdr = 4
    headers3 = ["№","Ген","Субтип (модел)"] + [SUBTYPE_LABEL[c] for c in cls_cols] + ["PAM50?","STRING"]
    for ci, h in enumerate(headers3, 1):
        ws3.cell(hdr, ci, h)
    style_header(ws3, hdr, len(headers3))

    morder = {st: i for i, st in enumerate(cls_cols)}
    genes_m = sorted(genes, key=lambda g: (morder.get(model_assign.get(g), 99), g))
    r = hdr + 1
    for i, g in enumerate(genes_m, 1):
        st = model_assign.get(g)
        color = SUBTYPE_COLORS.get(st, "BFBFBF")
        ws3.cell(r, 1, i).alignment = CENTER
        ws3.cell(r, 2, g)
        ws3.cell(r, 3, SUBTYPE_LABEL.get(st, st))
        for ci, c in enumerate(cls_cols):
            ws3.cell(r, 4 + ci, round(float(imp.loc[g, c]), 4)).alignment = CENTER
        col_pam = 4 + len(cls_cols)
        ws3.cell(r, col_pam, "ДА" if g in PAM50 else "—").alignment = CENTER
        ws3.cell(r, col_pam + 1, "Изолиран" if degree.get(g, 0) == 0 else "Свързан").alignment = CENTER
        for c in range(1, 4):
            cell = ws3.cell(r, c)
            cell.border = BORDER
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(color="FFFFFF", bold=(c == 2))
        for c in range(4, len(headers3) + 1):
            ws3.cell(r, c).border = BORDER
        r += 1
    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 18
    for ci in range(len(cls_cols)):
        ws3.column_dimensions[get_column_letter(4 + ci)].width = 12
    ws3.column_dimensions[get_column_letter(4 + len(cls_cols))].width = 9
    ws3.column_dimensions[get_column_letter(5 + len(cls_cols))].width = 11
    ws3.freeze_panes = "C5"

    # ---- TAB 4: сравнение ----
    ws4 = wb.create_sheet("Сравнение лит. vs модел")
    ws4.cell(1, 1, "Сравнение: субтип по ЛИТЕРАТУРА (Tab 2) срещу субтип по МОДЕЛА (Tab 3)").font = Font(bold=True, size=13)
    ws4.cell(2, 1, "Зелено = съвпадат; жълто = моделът намира роля там, където литературата няма; червено = разминаване.").font = Font(italic=True, size=10, color="666666")
    hdr = 4
    for ci, h in enumerate(["№","Ген","Литература","Модел","Съвпадение","PAM50?"], 1):
        ws4.cell(hdr, ci, h)
    style_header(ws4, hdr, 6)
    MATCH = PatternFill("solid", fgColor="C6EFCE"); MATCH_F = Font(color="006100")
    NEWROLE = PatternFill("solid", fgColor="FFEB9C"); NEWROLE_F = Font(color="9C6500")
    MISM = PatternFill("solid", fgColor="FFC7CE"); MISM_F = Font(color="9C0006")
    n_match = n_new = n_mis = 0
    r = hdr + 1
    for i, g in enumerate(sorted(genes), 1):
        lit = gene_subtype.get(g)          # None ако няма литературна роля
        mdl = model_assign.get(g)
        if lit is None:
            status, fill, font = "Само модел", NEWROLE, NEWROLE_F; n_new += 1
        elif lit == mdl:
            status, fill, font = "СЪВПАДА", MATCH, MATCH_F; n_match += 1
        else:
            status, fill, font = "Разминаване", MISM, MISM_F; n_mis += 1
        ws4.cell(r, 1, i).alignment = CENTER
        ws4.cell(r, 2, g).font = Font(bold=True)
        ws4.cell(r, 3, SUBTYPE_LABEL.get(lit, "—"))
        ws4.cell(r, 4, SUBTYPE_LABEL.get(mdl, mdl))
        ws4.cell(r, 5, status)
        ws4.cell(r, 6, "ДА" if g in PAM50 else "—").alignment = CENTER
        for c in range(1, 7):
            cell = ws4.cell(r, c); cell.border = BORDER
            if c == 5:
                cell.fill = fill; cell.font = font; cell.alignment = CENTER
        r += 1
    for col, w in zip("ABCDEF", [6,16,22,18,16,9]):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A5"
    ws4.cell(r + 1, 2, f"Съвпадат: {n_match}  |  Само модел (лит. няма роля): {n_new}  |  Разминавания: {n_mis}").font = Font(bold=True)
    print(f"  Tab3/4: модел-назначения готови | съвпадат={n_match}, само-модел={n_new}, разминавания={n_mis}")

    # ---- TAB 5: ВИЗУАЛИЗАЦИЯ (bar charts) ----
    from openpyxl.chart import BarChart, Reference
    ws5 = wb.create_sheet("Визуализация")
    ws5.cell(1, 1, "Визуализация на saliency на модела (best_model_tcga_196genes_09.06.pt)").font = Font(bold=True, size=13)

    imp2 = imp.copy()
    imp2['mean_sal'] = imp2[cls_cols].mean(axis=1)

    # --- Графика 1: Топ 15 гена по среден saliency ---
    top15 = imp2['mean_sal'].sort_values(ascending=False).head(15)
    ws5.cell(3, 1, "Топ 15 гена — среден saliency (глобална зависимост на модела)").font = Font(bold=True)
    ws5.cell(4, 1, "Ген"); ws5.cell(4, 2, "Среден saliency")
    for ci in (1, 2):
        ws5.cell(4, ci).font = Font(bold=True)
    for i, (g, v) in enumerate(top15.items(), start=5):
        ws5.cell(i, 1, g); ws5.cell(i, 2, round(float(v), 4))
    ch1 = BarChart(); ch1.type = "bar"; ch1.title = "Топ 15 гена по среден saliency"
    ch1.height = 9; ch1.width = 18; ch1.legend = None
    d1 = Reference(ws5, min_col=2, min_row=4, max_row=4 + len(top15))
    c1 = Reference(ws5, min_col=1, min_row=5, max_row=4 + len(top15))
    ch1.add_data(d1, titles_from_data=True); ch1.set_categories(c1)
    ws5.add_chart(ch1, "D3")

    # --- Графика 2: канонични маркери × 5 класа (subtype-специфичност?) ---
    canon = [g for g in ["ESR1","ERBB2","FOXA1","GATA3","TP53","MKI67","GRB7","FOXC1"] if g in imp2.index]
    base = 24
    ws5.cell(base, 1, "Канонични маркери — saliency по 5-те субтипа (плосък профил = НЕ е субтип-специфичен)").font = Font(bold=True)
    hb = base + 1
    ws5.cell(hb, 1, "Ген").font = Font(bold=True)
    for j, c in enumerate(cls_cols, start=2):
        ws5.cell(hb, j, SUBTYPE_LABEL[c]).font = Font(bold=True)
    for i, g in enumerate(canon, start=hb + 1):
        ws5.cell(i, 1, g)
        for j, c in enumerate(cls_cols, start=2):
            ws5.cell(i, j, round(float(imp2.loc[g, c]), 4))
    ch2 = BarChart(); ch2.type = "col"; ch2.grouping = "clustered"
    ch2.title = "Saliency на канонични маркери по субтип"
    ch2.height = 10; ch2.width = 22
    d2 = Reference(ws5, min_col=2, max_col=1 + len(cls_cols), min_row=hb, max_row=hb + len(canon))
    c2 = Reference(ws5, min_col=1, min_row=hb + 1, max_row=hb + len(canon))
    ch2.add_data(d2, titles_from_data=True); ch2.set_categories(c2)
    ws5.add_chart(ch2, "D24")
    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 16
    print("  Tab5: визуализации добавени")
else:
    print(f"  [ПРОПУСК] {IMP_CSV} липсва — Tab3/Tab4 не са създадени. Пусни model_gene_importance.py първо.")

OUT = "Gene_Table_196genes.xlsx"
try:
    wb.save(OUT)
except PermissionError:
    OUT = "Gene_Table_196genes_v2.xlsx"
    wb.save(OUT)
    print(f"[ВНИМАНИЕ] Оригиналният файл е отворен/заключен — записах в {OUT}")
print(f"\n[OK] Записан файл: {OUT}")
print(f"  Tab1: {len(genes)} гена ({len(pam50_present)} PAM50 червени най-отдолу)")
print(f"  Tab2: разпределение по субтип -> " +
      ", ".join(f"{SUBTYPE_LABEL[st]}={counts.get(st,0)}" for st in ["BRCA_Basal","BRCA_Her2","BRCA_LumA","BRCA_LumB","BRCA_Normal",None]))
